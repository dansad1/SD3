from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ExcelReader:

    HEADER_ROW_INDEX = 0
    DATA_START_INDEX = 3

    def __init__(
        self,
        max_rows=5000,
    ):
        self.max_rows = max_rows

    def read(
        self,
        file,
    ):
        try:
            file.seek(0)

            workbook = load_workbook(
                file,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except (
            InvalidFileException,
            OSError,
            ValueError,
            KeyError,
        ) as exc:
            raise ValidationError(
                "Не удалось прочитать Excel-файл",
            ) from exc

        sheet = workbook.active

        rows = list(
            sheet.iter_rows(
                values_only=True,
            )
        )

        if len(rows) <= self.DATA_START_INDEX:
            raise ValidationError(
                "Файл не содержит данных",
            )

        headers = [
            str(value).strip()
            if value is not None
            else ""
            for value in rows[
                self.HEADER_ROW_INDEX
            ]
        ]

        if not any(headers):
            raise ValidationError(
                "Файл не содержит заголовков",
            )

        data_rows = [
            row
            for row in rows[
                self.DATA_START_INDEX:
            ]
            if not self.is_empty_row(
                row,
            )
        ]

        if not data_rows:
            raise ValidationError(
                "Файл не содержит данных",
            )

        if len(data_rows) > self.max_rows:
            raise ValidationError(
                (
                    "Слишком много строк. "
                    f"Максимум: {self.max_rows}"
                ),
            )

        return {
            "headers": headers,
            "rows": data_rows,
            "start_row": (
                self.DATA_START_INDEX + 1
            ),
        }

    @staticmethod
    def is_empty_row(
        row,
    ):
        return all(
            value is None
            or (
                isinstance(
                    value,
                    str,
                )
                and not value.strip()
            )
            for value in row
        )