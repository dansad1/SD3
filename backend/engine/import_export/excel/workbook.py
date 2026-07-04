from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


class ExcelWorkbook:

    CONTENT_TYPE = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    LABEL_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    HINT_FILL = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
    )

    LABEL_FONT = Font(
        bold=True,
    )

    HINT_FONT = Font(
        italic=True,
        size=9,
        color="666666",
    )

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ALIGN = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    def __init__(
        self,
        title,
    ):
        self.workbook = Workbook()

        self.sheet = self.workbook.active
        self.sheet.title = title

    # =====================================================
    # HEADERS
    # =====================================================

    def write_headers(
        self,
        headers,
    ):
        self.sheet.append([
            header["name"]
            for header in headers
        ])

        self.sheet.append([
            header["label"]
            for header in headers
        ])

        self.sheet.append([
            header.get(
                "hint",
                "",
            )
            for header in headers
        ])

        self._style_headers()

    def _style_headers(
        self,
    ):
        for row in (
            1,
            2,
            3,
        ):
            for cell in self.sheet[row]:
                cell.border = self.BORDER
                cell.alignment = self.ALIGN

        for cell in self.sheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        for cell in self.sheet[2]:
            cell.fill = self.LABEL_FILL
            cell.font = self.LABEL_FONT

        for cell in self.sheet[3]:
            cell.fill = self.HINT_FILL
            cell.font = self.HINT_FONT

        self.sheet.freeze_panes = "A4"

    # =====================================================
    # ROWS
    # =====================================================

    def write_rows(
        self,
        rows,
    ):
        if not rows:
            self._finish()
            return

        columns = [
            cell.value
            for cell in self.sheet[1]
        ]

        for row in rows:

            self.sheet.append([
                row.get(
                    column,
                    "",
                )
                for column in columns
            ])

        self._finish()

    # =====================================================
    # FINISH
    # =====================================================

    def _finish(
        self,
    ):
        self.sheet.auto_filter.ref = (
            self.sheet.dimensions
        )

        for column in self.sheet.columns:

            width = 10

            for cell in column:

                if cell.value is None:
                    continue

                width = max(
                    width,
                    len(
                        str(
                            cell.value,
                        )
                    ) + 3,
                )

            width = min(
                width,
                40,
            )

            self.sheet.column_dimensions[
                column[0].column_letter
            ].width = width

    # =====================================================
    # RESPONSE
    # =====================================================

    def to_response(
        self,
        filename,
    ):
        buffer = BytesIO()

        self.workbook.save(
            buffer,
        )

        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type=self.CONTENT_TYPE,
        )

        response[
            "Content-Disposition"
        ] = (
            "attachment; "
            f'filename="{filename}"'
        )

        return response